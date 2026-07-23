"""
Daily "Benchmark Beaters" screener -> filled Excel workbook -> PDF.

Screens the TSX (250M+) universe against XIC.TO and the S&P 500 universe
against SPY for stocks making a 6-month relative high in the last 5 trading
days, tracks how many consecutive weeks each ticker has shown up (New /
Repeat / Streak), enriches the list with company data pulled from a Koyfin
screener export, and writes the results into the print-ready "Cdn Hardcoded"
/ "US Hardcoded" tabs of templates/benchmark_beaters_template.xlsx. The
workbook is then converted to PDF via headless LibreOffice.

The only manual step left is exporting the two Koyfin screener CSVs
(data/koyfin_cdn.csv, data/koyfin_us.csv) before the workflow runs.
"""

import os
import shutil
import subprocess
import sys
from copy import copy
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common_screening import load_tsx_symbols, load_sp500_symbols, screen_market

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(REPO_ROOT, "templates", "benchmark_beaters_template.xlsx")
KOYFIN_CDN_PATH = os.path.join(REPO_ROOT, "data", "koyfin_cdn.csv")
KOYFIN_US_PATH = os.path.join(REPO_ROOT, "data", "koyfin_us.csv")
OUTPUT_DIR = os.path.join(REPO_ROOT, "outputs", "benchmark-beaters")

NAME_SUFFIXES = ["Ltd.", "Inc.", "Corp."]

DATA_START_ROW = 8
DATA_END_COL = 9  # column I (Business Description) - last real data column
TABLE_COLS = list(range(2, DATA_END_COL + 1))  # B..I

CTA_TEXT = (
    "\U0001F513 Want to know which of these we would actually buy? "
    "Start your 14-Day Free Trial at www.5iresearch.ca/bb"
)
DISCLOSURE_TEXT = (
    "Disclosure: While this table does not constitute any formal opinion on names presented, "
    "authors may own shares in non-Canadian securities listed above."
)
ORANGE = "FFC67A29"
ORANGE_FILL = "FFFCEFE0"
LIGHT_GRAY_FILL = "FFF2F2F2"

SIGNAL_BADGE = {
    # (label, ARGB fill) - plain cell formatting, not emoji: emoji fall back to
    # mismatched monochrome glyphs (e.g. a flat blue box) under LibreOffice's
    # headless PDF renderer, which has no color-emoji font installed.
    "New": ("NEW", "FF2E86DE"),
    "Repeat": ("RPT", "FF8E44AD"),
    "Streak": ("STRK", "FFC0392B"),
}

# A fixed row height was too short for the longest wrapped Business
# Description / Company text -> LibreOffice rendered the overflow bleeding
# past the row's colored fill into the row below, which showed up as the
# left/right edge columns looking mismatched. Size each row to what its
# actual (longest) wrapped cell needs instead.
MIN_ROW_HEIGHT = 26
LINE_HEIGHT = 17     # deliberately generous - erring tall is a cosmetic
PADDING = 16          # nit, erring short reproduces the color-bleed bug
DESC_CHARS_PER_LINE = 68    # column I, width 56, 9.5pt font
COMPANY_CHARS_PER_LINE = 20  # column D, width 26, bold


def _row_height_for(company, description):
    import math
    desc_lines = math.ceil(len(description or "") / DESC_CHARS_PER_LINE) or 1
    company_lines = math.ceil(len(company or "") / COMPANY_CHARS_PER_LINE) or 1
    lines = max(desc_lines, company_lines, 1)
    return max(MIN_ROW_HEIGHT, lines * LINE_HEIGHT + PADDING)


def load_koyfin(path, market_label):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Missing Koyfin export for {market_label}: {path}\n"
            "Export the screener from Koyfin as CSV and commit it at that path before running."
        )
    df = pd.read_csv(path)
    df["Ticker"] = df["Ticker"].astype(str).str.strip().str.upper()
    return df.set_index("Ticker")


def short_description(text):
    if not isinstance(text, str) or not text:
        return ""
    normalized = text
    for suffix in NAME_SUFFIXES:
        normalized = normalized.replace(suffix, suffix.rstrip("."))
    return normalized.split(".")[0].strip()


def build_hardcoded_table(screen_df, koyfin_df, strip_suffix=None):
    records = []
    for _, row in screen_df.iterrows():
        raw_ticker = row["Ticker"]
        lookup_ticker = raw_ticker.replace(strip_suffix, "") if strip_suffix else raw_ticker
        if lookup_ticker not in koyfin_df.index:
            print(f"  no Koyfin match for {raw_ticker}, skipping")
            continue
        info = koyfin_df.loc[lookup_ticker]
        if isinstance(info, pd.DataFrame):
            info = info.iloc[0]
        records.append(
            {
                "Ticker": lookup_ticker,
                "Company": info.get("Name", ""),
                "Sector": info.get("Sector", ""),
                "Industry": info.get("Industry", ""),
                "6-Mo Return": round(row["6M_Perf_%"] / 100, 4),
                "Signal": row["Signal"],
                "Mkt Cap ($M)": info.get("Market Cap", ""),
                "Business Description": short_description(info.get("Description", "")),
            }
        )
    return pd.DataFrame(records)


def _copy_row_style(ws, dst_row, src_row):
    for col in TABLE_COLS:
        dst = ws.cell(row=dst_row, column=col)
        dst._style = copy(ws.cell(row=src_row, column=col)._style)


def _set_conditional_formatting_range(ws, last_data_row):
    """The color-scale on the Return column is scoped to a fixed range in the
    template; repoint it at the actual number of data rows each run so it
    never colors (or fails to color) rows that don't exist today."""
    to_replace = [(cf, cf.rules) for cf in list(ws.conditional_formatting) if str(cf.sqref).startswith("G8:G")]
    for cf, rules in to_replace:
        del ws.conditional_formatting._cf_rules[cf]
        for rule in rules:
            ws.conditional_formatting.add(f"G8:G{last_data_row}", rule)


def _write_banner(ws, row, height, text, fill_hex, font_color, bold, italic=False, border_hex=None):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=DATA_END_COL)
    cell = ws.cell(row=row, column=2, value=text)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(bold=bold, italic=italic, color=font_color, size=10)
    fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    border = Border(*(Side(style="thin", color=border_hex),) * 4) if border_hex else Border()
    for col in TABLE_COLS:
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.border = border


def write_hardcoded_sheet(ws, table_df):
    n = len(table_df)

    for offset in range(n):
        row = DATA_START_ROW + offset
        src_row = DATA_START_ROW if offset % 2 == 0 else DATA_START_ROW + 1
        if row != src_row:
            _copy_row_style(ws, row, src_row)
        record = table_df.iloc[offset]
        ws.row_dimensions[row].height = _row_height_for(record["Company"], record["Business Description"])
        label, badge_fill = SIGNAL_BADGE.get(record["Signal"], ("", None))
        badge_cell = ws.cell(row=row, column=2, value=label)  # B: signal badge
        if badge_fill:
            badge_cell.fill = PatternFill(fill_type="solid", fgColor=badge_fill)
            badge_cell.font = Font(bold=True, color="FFFFFFFF", size=7)
            badge_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3, value=record["Ticker"])
        ws.cell(row=row, column=4, value=record["Company"])
        ws.cell(row=row, column=5, value=record["Sector"])
        ws.cell(row=row, column=6, value=record["Industry"])
        ws.cell(row=row, column=7, value=record["6-Mo Return"])
        ws.cell(row=row, column=8, value=record["Mkt Cap ($M)"])
        ws.cell(row=row, column=9, value=record["Business Description"])

    if n == 0:
        last_data_row = DATA_START_ROW - 1
    else:
        last_data_row = DATA_START_ROW + n - 1
        # heavier bottom border under the last row, matching the template's
        # original "end of table" styling
        for col in TABLE_COLS:
            c = ws.cell(row=last_data_row, column=col)
            b = c.border
            c.border = Border(left=b.left, right=b.right, top=b.top, bottom=Side(style="medium"))

    _set_conditional_formatting_range(ws, max(last_data_row, DATA_START_ROW))

    # The template has pre-baked banding/fill for a wide fixed row range (a
    # leftover of the original hand-built file). Neutralize the spacer row
    # between the table and the CTA banner so no stray colored cell (e.g. the
    # Return column's base green fill) shows through from that inherited style.
    spacer_row = last_data_row + 1
    for col in TABLE_COLS:
        c = ws.cell(row=spacer_row, column=col)
        c.value = None
        c.fill = PatternFill(fill_type=None)
        c.border = Border()
        c.number_format = "General"
    ws.row_dimensions[spacer_row].height = 8

    cta_row = last_data_row + 2
    _write_banner(
        ws, cta_row, height=34, text=CTA_TEXT,
        fill_hex=ORANGE_FILL, font_color=ORANGE, bold=True, border_hex=ORANGE,
    )
    disclosure_row = cta_row + 1
    _write_banner(
        ws, disclosure_row, height=28, text=DISCLOSURE_TEXT,
        fill_hex=LIGHT_GRAY_FILL, font_color="FF666666", bold=False, italic=True,
    )

    ws.print_area = f"A1:J{disclosure_row}"


PRINT_SHEETS = ["Cdn Hardcoded", "US Hardcoded"]


def make_print_ready_copy(source_xlsx_path, dest_xlsx_path):
    """PDF export should only contain the two print-ready tabs, not the
    Instructions/Date/legacy manual-workflow tabs. LibreOffice's --convert-to
    exports every sheet in the workbook, so build a stripped-down copy first."""
    wb = load_workbook(source_xlsx_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in PRINT_SHEETS:
            del wb[sheet_name]
    wb.save(dest_xlsx_path)


def convert_to_pdf(xlsx_path, out_dir):
    subprocess.run(
        [
            "soffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            out_dir,
            xlsx_path,
        ],
        check=True,
        timeout=300,
    )
    pdf_path = os.path.join(out_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
    if not os.path.exists(pdf_path):
        raise RuntimeError(f"LibreOffice did not produce expected PDF at {pdf_path}")
    return pdf_path


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")

    tsx_symbols = load_tsx_symbols()
    cdn_screen = screen_market(tsx_symbols, "XIC.TO", "Cdn")

    sp500_symbols = load_sp500_symbols()
    us_screen = screen_market(sp500_symbols, "SPY", "US")

    koyfin_cdn = load_koyfin(KOYFIN_CDN_PATH, "Cdn")
    koyfin_us = load_koyfin(KOYFIN_US_PATH, "US")

    cdn_table = build_hardcoded_table(cdn_screen, koyfin_cdn, strip_suffix=".TO")
    us_table = build_hardcoded_table(us_screen, koyfin_us)

    wb = load_workbook(TEMPLATE_PATH)
    wb["Date"]["B2"] = today
    wb["Cdn Hardcoded"]["H2"] = today
    wb["US Hardcoded"]["H2"] = today
    write_hardcoded_sheet(wb["Cdn Hardcoded"], cdn_table)
    write_hardcoded_sheet(wb["US Hardcoded"], us_table)

    xlsx_out = os.path.join(OUTPUT_DIR, f"Benchmark_Beaters_{date_str}.xlsx")
    wb.save(xlsx_out)
    print(f"Saved workbook: {xlsx_out}")

    print_ready_path = os.path.join(OUTPUT_DIR, f"_print_ready_{date_str}.xlsx")
    make_print_ready_copy(xlsx_out, print_ready_path)
    pdf_out = convert_to_pdf(print_ready_path, OUTPUT_DIR)
    final_pdf = os.path.join(OUTPUT_DIR, f"Benchmark_Beaters_{date_str}.pdf")
    os.replace(pdf_out, final_pdf)
    os.remove(print_ready_path)
    print(f"Saved PDF: {final_pdf}")

    latest_xlsx = os.path.join(OUTPUT_DIR, "Benchmark_Beaters_latest.xlsx")
    wb.save(latest_xlsx)
    shutil.copyfile(final_pdf, os.path.join(OUTPUT_DIR, "Benchmark_Beaters_latest.pdf"))


if __name__ == "__main__":
    main()
